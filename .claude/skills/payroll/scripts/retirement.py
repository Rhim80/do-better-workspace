#!/usr/bin/env python3
"""
퇴직금 계산 + 퇴직금산정내역서 Excel 생성

Usage:
    python3 retirement.py 홍길동 --date 2026-02-28                    # 계산만
    python3 retirement.py 홍길동 --date 2026-02-28 --excel            # 계산 + Excel 생성
    python3 retirement.py 홍길동 --date 2026-02-28 --excel --save     # 계산 + Excel + 기록
    python3 retirement.py 홍길동 --date 2026-02-28 --wages 2666667 2400000 2442105
    python3 retirement.py --calc 83430 608                            # 직접 계산

퇴직금 공식:
    퇴직금 = 평균임금 × 30일 × (재직일수 / 365)
    평균임금 = 퇴직 전 3개월 총 급여 / 퇴직 전 3개월 총 일수

주의:
    Google Sheets 급여 데이터는 자체 입력이므로 세무사 확정 급여대장과 다를 수 있음.
    퇴직금 계산 시 반드시 세무사 확정 급여대장 첨부파일을 다운로드하여 대조 검증 필요.
    --wages 옵션으로 세무사 확정 수치를 직접 지정 가능.
"""

import sys
import argparse
from pathlib import Path
from datetime import datetime, date, timedelta
from typing import Dict, List, Optional, Tuple
import yaml
import json

sys.path.insert(0, str(Path.home() / '.claude/lib'))
sys.path.insert(0, str(Path(__file__).parent))
from google_auth import get_credentials
from googleapiclient.discovery import build
from config_loader import (
    load_config as _load_config,
    get_company_short,
    get_business_owner,
    resolve_path,
)

CONFIG_PATH = Path(__file__).parent.parent / 'config' / 'rates-kr-2026.yaml'


def load_config() -> Dict:
    """company.yaml + rates-kr-2026.yaml 통합 로드."""
    return _load_config()


def _legacy_load_config() -> Dict:
    if not CONFIG_PATH.exists():
        print(f"Error: 설정 파일이 없습니다: {CONFIG_PATH}")
        sys.exit(1)
    with open(CONFIG_PATH, 'r', encoding='utf-8') as f:
        return yaml.safe_load(f)


def get_sheets_service():
    creds = get_credentials()
    if not creds:
        print("Error: 인증 실패")
        sys.exit(1)
    return build('sheets', 'v4', credentials=creds)


def parse_date(date_str: str) -> Optional[date]:
    """날짜 문자열 파싱"""
    if not date_str:
        return None
    for fmt in ['%Y-%m-%d', '%Y.%m.%d', '%Y/%m/%d']:
        try:
            return datetime.strptime(date_str, fmt).date()
        except ValueError:
            continue
    return None


def days_in_month(year: int, month: int) -> int:
    """해당 월의 일수"""
    if month == 12:
        next_month = date(year + 1, 1, 1)
    else:
        next_month = date(year, month + 1, 1)
    return (next_month - date(year, month, 1)).days


def get_3month_periods(resignation_date: date) -> List[Tuple[str, str, str, int]]:
    """퇴직 전 3개월 기간 계산 (역순: 퇴직월, 전월, 전전월)
    Returns: [(year_month, start_date, end_date, days), ...]
    """
    periods = []
    for i in range(3):
        target = resignation_date - timedelta(days=i * 30)
        ym = target.strftime('%Y-%m')
        y, m = target.year, target.month
        d = days_in_month(y, m)
        start = f"{y}-{m:02d}-01"
        end = f"{y}-{m:02d}-{d:02d}"
        periods.append((ym, start, end, d))
    # 역순 → 시간순
    periods.reverse()
    return periods


# ============================================================
# Google Sheets 조회
# ============================================================

def get_employee_data(service, spreadsheet_id: str, employee_name: str) -> Optional[Dict]:
    """직원DB에서 직원 데이터 조회"""
    try:
        result = service.spreadsheets().values().get(
            spreadsheetId=spreadsheet_id, range='직원DB'
        ).execute()
        values = result.get('values', [])
        if len(values) < 2:
            return None
        headers = values[0]
        for row in values[1:]:
            emp = dict(zip(headers, row + [''] * (len(headers) - len(row))))
            if emp.get('이름', '') == employee_name:
                return emp
        return None
    except Exception as e:
        print(f"Error: {e}")
        return None


def get_recent_payroll(service, spreadsheet_id: str, employee_name: str,
                       resignation_date: date) -> Tuple[List[Dict], int]:
    """최근 3개월 급여 조회 (월별 상세)

    Returns:
        monthly_data: [{'year_month': '2026-01', 'total': 2400000, ...}, ...]
        total_days: 총 일수
    """
    monthly_data = []
    total_days = 0
    periods = get_3month_periods(resignation_date)

    for ym, start, end, days in periods:
        sheet_name = f"{ym}_급여대장"
        entry = {'year_month': ym, 'days': days, 'total': 0,
                 'base': 0, 'meal': 0, 'allowance': 0, 'found': False}

        try:
            result = service.spreadsheets().values().get(
                spreadsheetId=spreadsheet_id, range=sheet_name
            ).execute()
            values = result.get('values', [])
            if len(values) < 2:
                monthly_data.append(entry)
                total_days += days
                continue

            headers = values[0]
            for row in values[1:]:
                if len(row) > 0 and row[0] == employee_name:
                    emp = dict(zip(headers, row + [''] * (len(headers) - len(row))))
                    entry['found'] = True
                    entry['total'] = int(str(emp.get('지급총액', '0')).replace(',', ''))
                    entry['base'] = int(str(emp.get('기본급', '0')).replace(',', ''))
                    entry['meal'] = int(str(emp.get('식대', '0')).replace(',', ''))
                    # 추가수당 또는 시간외수당
                    allow = emp.get('추가수당', emp.get('시간외수당', '0'))
                    entry['allowance'] = int(str(allow).replace(',', ''))
                    break
        except Exception:
            pass

        monthly_data.append(entry)
        total_days += days

    return monthly_data, total_days


def update_employee_resign_date(service, spreadsheet_id: str,
                                employee_name: str, resign_date: str):
    """직원DB 퇴사일 업데이트"""
    result = service.spreadsheets().values().get(
        spreadsheetId=spreadsheet_id, range='직원DB'
    ).execute()
    values = result.get('values', [])
    headers = values[0]
    resign_idx = headers.index('퇴사일')

    for i, row in enumerate(values[1:], start=2):
        if len(row) > 0 and row[0] == employee_name:
            cell = f"직원DB!{chr(65 + resign_idx)}{i}"
            service.spreadsheets().values().update(
                spreadsheetId=spreadsheet_id, range=cell,
                valueInputOption='RAW',
                body={'values': [[resign_date]]}
            ).execute()
            print(f"  직원DB 퇴사일 업데이트: {employee_name} → {resign_date}")
            return True
    return False


# ============================================================
# 퇴직금 계산
# ============================================================

class RetirementCalculator:
    """퇴직금 계산기"""

    def __init__(self, config: Dict):
        self.config = config
        self.min_service_days = config['retirement']['min_service_days']

    def calculate_service_days(self, hire_date: date, resignation_date: date) -> int:
        """재직일수 계산 (입사일 포함)"""
        return (resignation_date - hire_date).days + 1

    def calculate(self, hire_date: date, resignation_date: date,
                  monthly_wages: List[int], total_days: int) -> Dict:
        """전체 퇴직금 계산"""
        service_days = self.calculate_service_days(hire_date, resignation_date)

        if service_days < self.min_service_days:
            return {
                'eligible': False,
                'reason': f'재직일수 {service_days}일 (최소 {self.min_service_days}일 필요)',
                'service_days': service_days
            }

        total_wage = sum(monthly_wages)
        avg_daily_wage = total_wage / total_days
        retirement_pay = int(avg_daily_wage * 30 * (service_days / 365))

        # 통상임금 비교 (기본급+식대 기준)
        ordinary_monthly = 2400000  # 기본적으로 기본급+식대
        ordinary_daily = ordinary_monthly * 3 / total_days

        return {
            'eligible': True,
            'hire_date': hire_date.isoformat(),
            'resignation_date': resignation_date.isoformat(),
            'service_days': service_days,
            'service_years': round(service_days / 365, 2),
            'monthly_wages': monthly_wages,
            'total_wage': total_wage,
            'total_days': total_days,
            'avg_daily_wage': round(avg_daily_wage, 2),
            'ordinary_daily_wage': round(ordinary_daily, 2),
            'applied_wage': 'average' if avg_daily_wage >= ordinary_daily else 'ordinary',
            'retirement_pay': retirement_pay
        }


# ============================================================
# 퇴직금산정내역서 Excel 생성
# ============================================================

def generate_retirement_excel(employee: Dict, result: Dict,
                              monthly_detail: List[Dict],
                              output_path: str,
                              cfg: Optional[Dict] = None) -> str:
    """퇴직금산정내역서 Excel 생성 (세무사 양식)

    Args:
        employee: 직원 정보 dict (이름, 주민번호, 매장, 입사일, 퇴사일 등)
        result: 퇴직금 계산 결과 dict
        monthly_detail: 월별 급여 상세 [{'year_month', 'base', 'meal', 'allowance', 'overtime', 'total', 'days'}, ...]
        output_path: 출력 파일 경로
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    except ImportError:
        print("Error: openpyxl이 필요합니다. pip install openpyxl")
        return ""

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "퇴직금산정내역서"

    # 스타일
    title_font = Font(name='맑은 고딕', size=16, bold=True)
    header_font = Font(name='맑은 고딕', size=11, bold=True)
    normal_font = Font(name='맑은 고딕', size=10)
    small_font = Font(name='맑은 고딕', size=9)
    center = Alignment(horizontal='center', vertical='center')
    left = Alignment(horizontal='left', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')

    for col, w in {'A': 16, 'B': 18, 'C': 16, 'D': 16, 'E': 16, 'F': 16, 'G': 14}.items():
        ws.column_dimensions[col].width = w

    def cell(row, col, value, font=normal_font, alignment=left, border=None, nf=None):
        c = ws.cell(row=row, column=col, value=value)
        c.font = font
        c.alignment = alignment
        if border:
            c.border = border
        if nf:
            c.number_format = nf
        return c

    def row_border(row, cols=7):
        for c in range(1, cols + 1):
            ws.cell(row=row, column=c).border = thin

    name = employee.get('이름', '')
    branch = employee.get('매장', '')
    ssn = employee.get('주민번호', '')
    phone = employee.get('연락처', '')
    hire = result['hire_date']
    resign = result['resignation_date']

    # === 제목 ===
    ws.merge_cells('A1:G1')
    cell(1, 1, "퇴직금산정내역서", title_font, center)

    # === 1. 인적사항 ===
    cell(3, 1, "1. 인적사항", header_font)

    company_short = get_company_short(cfg) if cfg else ''
    business_owner = get_business_owner(cfg) if cfg else ''
    company_label = f"{company_short} {branch}".strip() if branch else company_short

    info_rows = [
        (4, "성 명 :", name, "주민등록번호 :", ssn),
        (5, "주 소 :", "", "전 화 번 호 :", phone),
        (6, "입 사 일 :", hire, "퇴 사 일 :", resign),
        (7, "회 사 명 :", company_label, "사 업 주 명 :", business_owner),
    ]
    for r, l1, v1, l2, v2 in info_rows:
        cell(r, 1, l1, header_font, border=thin)
        cell(r, 2, v1, normal_font, border=thin)
        cell(r, 3, "", border=thin)
        cell(r, 4, "", border=thin)
        cell(r, 5, l2, header_font, border=thin)
        cell(r, 6, v2, normal_font, border=thin)
        cell(r, 7, "", border=thin)

    # === 2. 평균임금산정 ===
    cell(9, 1, "2. 평균임금산정", header_font)

    # 기간 헤더
    periods = monthly_detail
    r = 10
    cell(r, 1, "임 금 계 산", header_font, center, thin, None)
    ws.cell(r, 1).fill = fill
    cell(r, 2, "기간(부터)", normal_font, center, thin)
    ws.cell(r, 2).fill = fill
    for i, p in enumerate(periods):
        cell(r, 3 + i, f"{p['year_month']}-01", normal_font, center, thin)
    cell(r, 6, "", border=thin)

    r = 11
    cell(r, 1, "", border=thin)
    cell(r, 2, "기간(까지)", normal_font, center, thin)
    ws.cell(r, 2).fill = fill
    for i, p in enumerate(periods):
        ym = p['year_month']
        y, m = int(ym[:4]), int(ym[5:7])
        d = days_in_month(y, m)
        cell(r, 3 + i, f"{ym}-{d:02d}", normal_font, center, thin)
    cell(r, 6, "", border=thin)

    r = 12
    cell(r, 1, "", border=thin)
    cell(r, 2, "총근로일수(일)", normal_font, center, thin)
    ws.cell(r, 2).fill = fill
    total_days = 0
    for i, p in enumerate(periods):
        cell(r, 3 + i, p['days'], normal_font, center, thin)
        total_days += p['days']
    cell(r, 6, total_days, header_font, center, thin)

    # 급여 항목
    side_labels = ["평", "균", "임", "금", "산", "정"]
    item_names = ["기 본 급", "연 장 근 로", "기 타 수 당", "식 대", "", ""]

    for j, (sl, iname) in enumerate(zip(side_labels, item_names)):
        r = 13 + j
        cell(r, 1, sl, header_font, center, thin)
        cell(r, 2, iname, normal_font, center, thin)

        row_total = 0
        for i, p in enumerate(periods):
            val = None
            if iname == "기 본 급":
                val = p.get('base', 0) or None
            elif iname == "연 장 근 로":
                val = p.get('overtime', 0) or None
            elif iname == "기 타 수 당":
                val = p.get('extra', 0) or None
            elif iname == "식 대":
                val = p.get('meal', 0) or None

            if val and val > 0:
                cell(r, 3 + i, val, normal_font, right_align, thin, '#,##0')
                row_total += val
            else:
                cell(r, 3 + i, "-", normal_font, center, thin)

        if row_total > 0:
            cell(r, 6, row_total, normal_font, right_align, thin, '#,##0')
        else:
            cell(r, 6, "", border=thin)

    # 합계
    r = 19
    cell(r, 1, "", border=thin)
    cell(r, 2, "합 계 액", header_font, center, thin)
    ws.cell(r, 2).fill = fill
    for i, p in enumerate(periods):
        cell(r, 3 + i, p['total'], header_font, right_align, thin, '#,##0')
    cell(r, 6, result['total_wage'], header_font, right_align, thin, '#,##0')

    # 상여금
    r = 20
    cell(r, 1, "상여금", normal_font, left, thin)
    cell(r, 2, "이전 1년간 받은 상여금 총액 =", small_font, left, thin)
    cell(r, 3, "", border=thin)
    cell(r, 4, "-", normal_font, center, thin)
    cell(r, 5, "× 3 / 12 =", small_font, center, thin)
    cell(r, 6, "-", normal_font, center, thin)

    # 평균임금
    r = 21
    cell(r, 1, "3월간 지급된 임금총액:", header_font, left, thin)
    cell(r, 2, "", border=thin)
    cell(r, 3, result['total_wage'], header_font, right_align, thin, '#,##0')
    cell(r, 4, "3월간 총일수 :", normal_font, center, thin)
    cell(r, 5, result['total_days'], header_font, center, thin)
    cell(r, 6, "평균임금 :", normal_font, center, thin)
    cell(r, 7, result['avg_daily_wage'], header_font, right_align, thin, '#,##0.00')

    # === 3. 근로년수 ===
    r = 23
    cell(r, 1, "3. 근로년수", header_font)
    r = 24
    cell(r, 1, "입 사 일 :", normal_font, left, thin)
    cell(r, 2, hire, normal_font, left, thin)
    cell(r, 3, "", border=thin)
    cell(r, 4, "퇴 사 일 :", normal_font, center, thin)
    cell(r, 5, resign, normal_font, left, thin)
    cell(r, 6, "근로년수 :", normal_font, center, thin)
    cell(r, 7, result['service_years'], header_font, right_align, thin)

    # === 4. 퇴직금산정 ===
    r = 26
    cell(r, 1, "4. 퇴직금산정", header_font)
    r = 27
    for i, lbl in enumerate(["일평균임금", "근속일수", "1년에 30일분", "", "퇴직금", "퇴직소득세"]):
        cell(r, 1 + i, lbl, header_font, center, thin)
        ws.cell(r, 1 + i).fill = fill

    r = 28
    cell(r, 1, result['avg_daily_wage'], normal_font, right_align, thin, '#,##0.00')
    cell(r, 2, f"× {result['service_days']}", normal_font, center, thin)
    cell(r, 3, "× 30 일분 / 365 일 =", small_font, center, thin)
    cell(r, 4, "", border=thin)
    cell(r, 5, result['retirement_pay'], header_font, right_align, thin, '#,##0')
    cell(r, 6, "", border=thin)  # 퇴직소득세 - 세무사 확정

    # 차감
    r = 30
    cell(r, 1, "출국만기보험", normal_font, left, thin)
    cell(r, 2, "", border=thin)
    cell(r, 3, "근로소득세주민세환급금", small_font, left, thin)
    cell(r, 4, "", border=thin)
    cell(r, 5, "차감잔액", header_font, center, thin)
    ws.cell(r, 5).fill = fill
    cell(r, 6, "", border=thin)

    # 서명
    cell(33, 1, f"{resign[:4]}년    {int(resign[5:7])}월    일", normal_font)
    cell(35, 1, "위의 사실을 확인합니다.", normal_font)
    cell(37, 4, "사 업 주 :", normal_font)
    cell(37, 5, business_owner, normal_font)
    cell(37, 6, "(인)", normal_font)
    cell(38, 4, "근 로 자 :", normal_font)
    cell(38, 5, name, normal_font)
    cell(38, 6, "(인)", normal_font)

    # 인쇄 설정
    ws.print_area = 'A1:G38'
    ws.page_setup.orientation = 'portrait'

    wb.save(output_path)
    return output_path


# ============================================================
# 메인 커맨드
# ============================================================

def cmd_calculate(args):
    """직원 퇴직금 계산 (메인 명령)"""
    config = load_config()
    service = get_sheets_service()
    spreadsheet_id = config['google_sheets'].get('spreadsheet_id', '')

    if not spreadsheet_id:
        print("Error: spreadsheet_id가 설정되지 않았습니다.")
        return

    # 직원 데이터 조회
    emp = get_employee_data(service, spreadsheet_id, args.name)
    if not emp:
        print(f"Error: 직원을 찾을 수 없습니다: {args.name}")
        return

    hire_date = parse_date(emp.get('입사일', ''))
    if not hire_date:
        print("Error: 입사일이 없습니다.")
        return

    resignation_date = parse_date(args.date) if args.date else date.today()
    if not resignation_date:
        print("Error: 퇴사일 형식이 잘못되었습니다.")
        return

    print(f"\n=== {args.name} 퇴직금 계산 ===\n")
    print(f"입사일: {hire_date}")
    print(f"퇴사일: {resignation_date}")
    print(f"매장: {emp.get('매장', '')}")

    # 최근 3개월 급여 조회
    monthly_data, total_days = get_recent_payroll(
        service, spreadsheet_id, args.name, resignation_date)

    # --wages 옵션으로 세무사 확정 수치 직접 지정
    if args.wages:
        print(f"\n[세무사 확정 급여 직접 지정]")
        for i, wage in enumerate(args.wages):
            if i < len(monthly_data):
                old = monthly_data[i]['total']
                monthly_data[i]['total'] = wage
                if old != wage:
                    print(f"  {monthly_data[i]['year_month']}: {old:,} → {wage:,} (차이: {wage - old:+,})")

    wages = [m['total'] for m in monthly_data]
    print(f"\n최근 3개월 급여:")
    for m in monthly_data:
        src = "확정" if m['found'] else "미확인"
        print(f"  {m['year_month']}: {m['total']:,}원 ({m['days']}일) [{src}]")
    print(f"  합계: {sum(wages):,}원 / {total_days}일")

    if not any(wages):
        print("\nError: 급여 데이터가 없습니다.")
        return

    # 계산
    calculator = RetirementCalculator(config)
    result = calculator.calculate(hire_date, resignation_date, wages, total_days)

    if not result['eligible']:
        print(f"\n퇴직금 미해당: {result['reason']}")
        return

    print(f"\n--- 계산 결과 ---")
    print(f"재직일수: {result['service_days']}일 ({result['service_years']}년)")
    print(f"평균임금(일): {result['avg_daily_wage']:,}원")
    print(f"통상임금(일): {result['ordinary_daily_wage']:,}원")
    print(f"적용: {'평균임금' if result['applied_wage'] == 'average' else '통상임금'}")
    print(f"\n퇴직금: {result['retirement_pay']:,}원")

    if args.json:
        print("\n--- JSON ---")
        print(json.dumps(result, indent=2, ensure_ascii=False))

    # Excel 생성
    if args.excel:
        # 월별 상세 데이터 구성
        monthly_detail = []
        for m in monthly_data:
            detail = {
                'year_month': m['year_month'],
                'days': m['days'],
                'base': m.get('base', 0),
                'overtime': m.get('allowance', 0) if m.get('allowance', 0) > 100000 else 0,
                'extra': m.get('allowance', 0) if 0 < m.get('allowance', 0) <= 100000 else 0,
                'meal': m.get('meal', 0),
                'total': m['total'],
            }
            monthly_detail.append(detail)

        year = resignation_date.strftime('%Y')
        output_dir = resolve_path(config, 'excel_folder', year)
        output_dir.mkdir(parents=True, exist_ok=True)
        company_short = get_company_short(config)
        output_path = output_dir / f"{company_short} {args.name} 퇴직금산정내역서.xlsx"

        generate_retirement_excel(emp, result, monthly_detail, str(output_path), cfg=config)
        print(f"\nExcel 생성: {output_path}")

    # PKM 기록 저장
    if args.save:
        save_retirement_record(args.name, emp, result, monthly_data, config)

    # 직원DB 퇴사일 업데이트
    if args.update_db:
        update_employee_resign_date(
            service, spreadsheet_id, args.name, resignation_date.isoformat())


def cmd_direct_calc(args):
    """직접 계산 (평균임금, 재직일수)"""
    avg_wage = args.calc[0]
    service_days = args.calc[1]
    retirement_pay = int(avg_wage * 30 * (service_days / 365))

    print(f"\n=== 퇴직금 계산 (직접 입력) ===\n")
    print(f"평균임금(일): {avg_wage:,}원")
    print(f"재직일수: {service_days}일")
    print(f"\n계산식: {avg_wage:,} x 30 x ({service_days} / 365)")
    print(f"퇴직금: {retirement_pay:,}원")


def save_retirement_record(name: str, emp: Dict, result: Dict,
                           monthly_data: List[Dict],
                           cfg: Optional[Dict] = None):
    """퇴직금 계산 기록 저장"""
    year = result['resignation_date'][:4]
    folder = resolve_path(cfg or {}, 'resignation_folder', year)
    folder.mkdir(parents=True, exist_ok=True)

    company_short = get_company_short(cfg) if cfg else ''
    accountant_company = (cfg or {}).get('accountant', {}).get('company', '세무사')
    branch_label = f"{company_short} {emp.get('매장', '')}".strip() if emp.get('매장') else company_short

    # 월별 급여 테이블
    wage_rows = ""
    for m in monthly_data:
        wage_rows += f"| {m['year_month']} ({m['days']}일) | {m['total']:,}원 |\n"

    content = f"""# {name} 퇴직금 계산서

## 1. 퇴사자 인적사항

| 항목 | 내용 |
|------|------|
| 성명 | {name} |
| 주민등록번호 | {emp.get('주민번호', '')} |
| 매장 | {branch_label} |
| 직급 | {emp.get('직급', '')} |
| 입사일 | {result['hire_date']} |
| 퇴사일 | {result['resignation_date']} |
| 재직일수 | {result['service_days']}일 ({result['service_years']}년) |

---

## 2. 최근 3개월 급여 내역

| 월 | 지급총액 |
|----|----------|
{wage_rows}| **합계 ({result['total_days']}일)** | **{result['total_wage']:,}원** |

---

## 3. 퇴직금 계산

| 항목 | 금액 |
|------|------|
| 3개월 총 급여 | {result['total_wage']:,}원 |
| 3개월 총 일수 | {result['total_days']}일 |
| 일 평균임금 | {result['avg_daily_wage']:,}원 |

```
퇴직금 = {result['avg_daily_wage']:,} x 30 x ({result['service_days']} / 365)
       = {result['retirement_pay']:,}원
```

| 항목 | 금액 |
|------|------|
| **퇴직금** | **{result['retirement_pay']:,}원** |

---

*계산일: {date.today().isoformat()}*
*최종 퇴직금은 세무사({accountant_company}) 확정 기준*
"""

    filepath = folder / f"{result['resignation_date']}_{name}_퇴직금계산.md"
    filepath.write_text(content, encoding='utf-8')
    print(f"\nPKM 기록: {filepath}")


def main():
    parser = argparse.ArgumentParser(
        description='퇴직금 계산 + 퇴직금산정내역서 Excel 생성',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
예시:
  python3 retirement.py 홍길동 --date 2026-02-28
  python3 retirement.py 홍길동 --date 2026-02-28 --wages 2666667 2400000 2442105 --excel --save
  python3 retirement.py --calc 83430 608
        """)

    parser.add_argument('name', nargs='?', help='직원명')
    parser.add_argument('--date', help='퇴사일 (YYYY-MM-DD)')
    parser.add_argument('--wages', nargs=3, type=int, metavar=('M1', 'M2', 'M3'),
                        help='세무사 확정 3개월 급여 (시간순: 전전월 전월 퇴직월)')
    parser.add_argument('--excel', action='store_true',
                        help='퇴직금산정내역서 Excel 생성')
    parser.add_argument('--save', action='store_true',
                        help='PKM에 계산 기록 저장')
    parser.add_argument('--update-db', action='store_true',
                        help='직원DB 퇴사일 업데이트')
    parser.add_argument('--json', action='store_true', help='JSON 출력')
    parser.add_argument('--calc', nargs=2, type=int, metavar=('WAGE', 'DAYS'),
                        help='직접 계산 (일평균임금, 재직일수)')

    args = parser.parse_args()

    if args.calc:
        cmd_direct_calc(args)
    elif args.name:
        cmd_calculate(args)
    else:
        parser.print_help()


if __name__ == '__main__':
    main()
