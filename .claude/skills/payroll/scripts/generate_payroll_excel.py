#!/usr/bin/env python3
"""
급여대장 Excel 생성 스크립트
- Google Sheets에서 직원 기본정보 읽기
- 급여변동.md 파싱하여 수당/일할계산/파트타이머 반영
- 세무사 전달용 Excel 생성 (매장별 시트)

Usage:
    python3 generate_payroll_excel.py 2026-02
    python3 generate_payroll_excel.py 2026-02 --preview
"""

import sys
import argparse
import subprocess
import re
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional, Any, Tuple
import yaml

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
except ImportError:
    print("Error: openpyxl 패키지가 필요합니다.")
    print("설치: pip install openpyxl")
    sys.exit(1)

sys.path.insert(0, str(Path.home() / '.claude/lib'))
sys.path.insert(0, str(Path(__file__).parent))
from google_auth import get_credentials
from googleapiclient.discovery import build
from config_loader import (
    load_config as _load_config,
    get_branches,
    get_branch_prefix,
    is_single_site,
    resolve_path,
)

# 모듈 전역에 캐싱되는 cfg (load_config 호출 후 사용 가능)
_CFG: Dict = {}

# 매장 설정 (load_config 호출 시 company.yaml에서 동적 채워짐)
BRANCH_ORDER: List[str] = []
BRANCH_KEYWORDS: Dict[str, str] = {}
PKM_BASE: Path = Path.cwd()

# 스타일 정의
HEADER_FONT = Font(bold=True, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
SECTION_FILL = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
SUBTOTAL_FILL = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
TOTAL_FILL = PatternFill(start_color='D9EAD3', end_color='D9EAD3', fill_type='solid')
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
MONEY_FORMAT = '#,##0"원"'


def load_config() -> Dict:
    """설정 파일 로드 + 매장 정보를 모듈 전역에 반영"""
    global BRANCH_ORDER, BRANCH_KEYWORDS, PKM_BASE, _CFG
    cfg = _load_config()
    _CFG = cfg
    BRANCH_ORDER = get_branches(cfg)
    BRANCH_KEYWORDS = get_branch_prefix(cfg)
    if is_single_site(cfg):
        # 단일 사업장: 매장 분리 없이 회사명을 단일 그룹으로
        company_name = cfg.get('company', {}).get('name', '본사')
        BRANCH_ORDER = [company_name]
        BRANCH_KEYWORDS = {company_name: company_name}
    output_base = cfg.get('output', {}).get('pkm_base', '')
    if output_base:
        expanded = Path(output_base).expanduser()
        PKM_BASE = expanded if expanded.is_absolute() else Path.cwd() / expanded
    else:
        PKM_BASE = Path.cwd()
    return cfg


def parse_money(text: str) -> int:
    """금액 문자열을 정수로 변환 (예: '2,400,000원' -> 2400000)"""
    if not text:
        return 0
    cleaned = re.sub(r'[^\d]', '', str(text))
    return int(cleaned) if cleaned else 0


def parse_hours(text: str) -> float:
    """시간 문자열을 숫자로 변환 (예: '8시간' -> 8.0, '4.5h' -> 4.5)"""
    if not text:
        return 0.0
    match = re.search(r'(\d+(?:\.\d+)?)', str(text))
    return float(match.group(1)) if match else 0.0


def _parse_markdown_table(content: str) -> List[List[str]]:
    """마크다운 테이블을 파싱하여 행 리스트로 반환"""
    rows = []
    lines = content.strip().split('\n')

    for line in lines:
        line = line.strip()
        if not line or line.startswith('|--') or line.startswith('|-'):
            continue
        if line.startswith('|') and line.endswith('|'):
            cells = [cell.strip() for cell in line.split('|')[1:-1]]
            if cells and cells[0] in ['직원명', '이름', '항목']:
                continue
            if cells:
                rows.append(cells)

    return rows


def parse_payroll_changes(year_month: str) -> Tuple[Dict, Dict, Dict, Dict]:
    """
    급여변동.md 파일에서 변동사항 파싱

    Returns:
        allowances: {이름: {amount, note, prorated, prorated_base, prorated_meal}}
        allowance_sections: 수당 상세 데이터 (섹션별)
        parttimer_sections: 파트타이머 데이터 (섹션별)
        daily_worker_sections: 일용직 데이터 (매장별)
    """
    year = year_month[:4]
    record_dir = resolve_path(_CFG, 'payroll_record', year)
    changes_path = record_dir / f"{year_month}_급여변동.md"

    if not changes_path.exists():
        print(f"Warning: 급여변동 파일 없음: {changes_path}")
        return {}, {}, {}, {}

    print(f"급여변동 파일 로드: {changes_path}")

    with open(changes_path, 'r', encoding='utf-8') as f:
        content = f.read()

    allowances = {}
    allowance_sections = {}
    parttimer_sections = {}
    daily_worker_sections = {}

    branch_pattern = r'##\s+\d+\.\s+(.+?)(?=\n##\s+\d+\.|\n##\s+요약|$)'
    branch_sections = re.findall(branch_pattern, content, re.DOTALL)

    for section in branch_sections:
        lines = section.strip().split('\n')
        branch_name = lines[0].strip() if lines else "기타"
        section_content = '\n'.join(lines[1:])

        # 정직원 휴일근무 수당 파싱
        holiday_match = re.search(
            r'###\s+정직원\s*-\s*휴일근무\s+수당.*?\n(.*?)(?=\n###|\n---|\Z)',
            section_content, re.DOTALL
        )
        if holiday_match:
            table_content = holiday_match.group(1)
            rows = _parse_markdown_table(table_content)

            section_key = f"{branch_name} - 휴일근무 수당"
            allowance_sections[section_key] = [
                {'headers': ['직원명', '기본급', '통상시급', '휴일근무', '가산율', '수당']}
            ]

            for row in rows:
                if len(row) >= 6 and not row[0].startswith('**'):
                    name = row[0].strip()
                    amount = parse_money(row[5])
                    if name and amount > 0:
                        if name not in allowances:
                            allowances[name] = {'amount': 0, 'note': ''}
                        allowances[name]['amount'] += amount
                        allowances[name]['note'] = '휴일근무'
                        allowance_sections[section_key].append({
                            'data': [name, row[1], row[2], row[3], row[4], amount]
                        })

        # 정직원 추가근무 수당 파싱
        extra_match = re.search(
            r'###\s+정직원\s*-\s*추가근무\s+수당.*?\n(.*?)(?=\n###|\n---|\Z)',
            section_content, re.DOTALL
        )
        if extra_match:
            table_content = extra_match.group(1)
            rows = _parse_markdown_table(table_content)

            section_key = f"{branch_name} - 추가근무 수당"
            allowance_sections[section_key] = [
                {'headers': ['직원명', '기본급', '통상시급', '추가근무', '가산율', '수당']}
            ]

            for row in rows:
                if len(row) >= 6 and not row[0].startswith('**'):
                    name = row[0].strip()
                    amount = parse_money(row[5])
                    if name and amount > 0:
                        if name not in allowances:
                            allowances[name] = {'amount': 0, 'note': ''}
                        allowances[name]['amount'] += amount
                        if allowances[name]['note']:
                            allowances[name]['note'] += ', 추가근무'
                        else:
                            allowances[name]['note'] = '추가근무'
                        allowance_sections[section_key].append({
                            'data': [name, row[1], row[2], row[3], row[4], amount]
                        })

        # 정직원 일할계산 파싱
        prorated_match = re.search(
            r'###\s+정직원\s*-\s*일할계산\s*\((.+?)\).*?\n(.*?)(?=\n###|\n---|\Z)',
            section_content, re.DOTALL
        )
        if prorated_match:
            emp_name = prorated_match.group(1).strip()
            table_content = prorated_match.group(2)

            days_match = re.search(r'근무일수.*?(\d+)일', table_content)
            total_match = re.search(r'총일수.*?(\d+)일', table_content)
            work_days = int(days_match.group(1)) if days_match else 0
            total_days = int(total_match.group(1)) if total_match else 31

            base_match = re.search(r'기본급.*?(\d[\d,]+)원.*?(\d[\d,]+)원', table_content)
            meal_match = re.search(r'식대.*?(\d[\d,]+)원.*?(\d[\d,]+)원', table_content)

            prorated_base = parse_money(base_match.group(2)) if base_match else 0
            prorated_meal = parse_money(meal_match.group(2)) if meal_match else 0

            if emp_name and (prorated_base > 0 or prorated_meal > 0):
                allowances[emp_name] = {
                    'amount': 0,
                    'note': f'일할계산 ({work_days}/{total_days}일)',
                    'prorated': True,
                    'prorated_base': prorated_base,
                    'prorated_meal': prorated_meal
                }

                section_key = f"{branch_name} - 일할계산 ({emp_name})"
                allowance_sections[section_key] = [
                    {'headers': ['항목', '전액', '일할계산']},
                    {'data': ['근무일수', f'{total_days}일', f'{work_days}일']},
                    {'data': ['기본급', base_match.group(1) + '원' if base_match else '-', f'{prorated_base:,}원']},
                    {'data': ['식대', meal_match.group(1) + '원' if meal_match else '-', f'{prorated_meal:,}원']},
                    {'data': ['합계', '-', f'{prorated_base + prorated_meal:,}원']}
                ]

        # 파트타이머 파싱 (일반)
        parttimer_match = re.search(
            r'###\s+파트타이머\s*\n(.*?)(?=\n###|\n---|\Z)',
            section_content, re.DOTALL
        )
        if parttimer_match:
            table_content = parttimer_match.group(1)
            rows = _parse_markdown_table(table_content)

            section_key = f"{branch_name} - 파트타이머"
            parttimer_sections[section_key] = [
                {'headers': ['직원명', '주민번호', '근무시간', '시급', '급여', '비고']}
            ]

            for row in rows:
                if len(row) >= 5 and not row[0].startswith('**'):
                    parttimer_sections[section_key].append({
                        'data': [
                            row[0],
                            row[1] if len(row) > 1 else '-',
                            row[2] if len(row) > 2 else '-',
                            row[3] if len(row) > 3 else '-',
                            parse_money(row[4]) if len(row) > 4 else 0,
                            row[5] if len(row) > 5 else '-'
                        ]
                    })

        # 파트타이머 주휴수당 추가지급 파싱
        bonus_match = re.search(
            r'###\s+파트타이머\s*-\s*(.+?주휴수당.+?)\n(.*?)(?=\n###|\n---|\Z)',
            section_content, re.DOTALL
        )
        if bonus_match:
            title = bonus_match.group(1).strip()
            table_content = bonus_match.group(2)
            rows = _parse_markdown_table(table_content)

            section_key = f"{branch_name} - {title}"
            parttimer_sections[section_key] = [
                {'headers': ['직원명', '주민번호', '11월분', '12월분', '1월분', '총 지급액']}
            ]

            for row in rows:
                if len(row) >= 5 and not row[0].startswith('**'):
                    parttimer_sections[section_key].append({
                        'data': [
                            row[0],
                            row[1] if len(row) > 1 else '-',
                            parse_money(row[2]) if len(row) > 2 else 0,
                            parse_money(row[3]) if len(row) > 3 else 0,
                            parse_money(row[4]) if len(row) > 4 else 0,
                            parse_money(row[5]) if len(row) > 5 else 0
                        ]
                    })

        # 일용직 파싱
        daily_match = re.search(
            r'###\s+일용직\s*\n(.*?)(?=\n###|\n---|\Z)',
            section_content, re.DOTALL
        )
        if daily_match:
            table_content = daily_match.group(1)
            rows = _parse_markdown_table(table_content)

            section_key = f"{branch_name} - 일용직"
            daily_worker_sections[section_key] = [
                {'headers': ['직원명', '주민번호', '지급액', '실제 근무일', '근무일수']}
            ]

            for row in rows:
                if len(row) >= 3 and not row[0].startswith('**'):
                    daily_worker_sections[section_key].append({
                        'data': [
                            row[0],
                            row[1] if len(row) > 1 else '-',
                            parse_money(row[2]) if len(row) > 2 else 0,
                            row[3] if len(row) > 3 else '-',
                            row[4] if len(row) > 4 else '-'
                        ]
                    })

    print(f"  - 정직원 수당/일할계산: {len(allowances)}명")
    print(f"  - 정직원_수당 섹션: {len(allowance_sections)}개")
    print(f"  - 파트타이머 섹션: {len(parttimer_sections)}개")
    print(f"  - 일용직 섹션: {len(daily_worker_sections)}개")

    return allowances, allowance_sections, parttimer_sections, daily_worker_sections


# === Google Sheets ===

def get_sheets_service():
    """Google Sheets API 서비스 생성"""
    creds = get_credentials()
    if not creds:
        print("Error: 인증 실패")
        sys.exit(1)
    return build('sheets', 'v4', credentials=creds)


def read_employees(service, spreadsheet_id: str) -> List[Dict]:
    """Google Sheets에서 직원 정보 읽기"""
    try:
        result = service.spreadsheets().values().get(
            spreadsheetId=spreadsheet_id,
            range='직원DB'
        ).execute()
        values = result.get('values', [])

        if len(values) < 2:
            return []

        headers = values[0]
        employees = []

        for row in values[1:]:
            if len(row) < 6:
                continue

            emp = {}
            for i, h in enumerate(headers):
                emp[h] = row[i] if i < len(row) else ''

            # 퇴사자 제외 (퇴사일이 있고 과거인 경우)
            if emp.get('퇴사일'):
                try:
                    quit_date = datetime.strptime(emp['퇴사일'], '%Y-%m-%d')
                    if quit_date < datetime.now():
                        if quit_date.month != datetime.now().month:
                            continue
                except:
                    pass

            try:
                emp['기본급'] = int(str(emp.get('기본급', 0)).replace(',', ''))
            except:
                emp['기본급'] = 0

            try:
                emp['식대'] = int(str(emp.get('식대', 200000)).replace(',', ''))
            except:
                emp['식대'] = 200000

            employees.append(emp)

        return employees

    except Exception as e:
        print(f"Error: 직원 정보 읽기 실패: {e}")
        return []


# === 급여 계산 ===

def calculate_payroll(base: int, meal: int = 200000, extra: int = 0) -> Dict:
    """payroll.py CLI 호출하여 공제 계산"""
    total_taxable = base + extra
    cmd = f"python3 {Path(__file__).parent}/payroll.py calculate {total_taxable} --meal {meal}"

    try:
        result = subprocess.run(cmd, shell=True, capture_output=True, text=True, timeout=30)
        lines = result.stdout.strip().split('\n')

        data = {
            '지급총액': base + meal + extra,
            '국민연금': 0, '건강보험': 0, '장기요양': 0,
            '고용보험': 0, '소득세': 0, '지방소득세': 0,
            '공제총액': 0, '실지급액': base + meal + extra
        }

        for line in lines:
            if ':' in line and '원' in line:
                parts = line.split(':')
                key = parts[0].strip()
                val = parts[1].strip().replace('원', '').replace(',', '')
                try:
                    data[key] = int(val)
                except:
                    pass

        return data

    except Exception as e:
        print(f"Warning: 급여 계산 실패: {e}")
        return {
            '지급총액': base + meal + extra,
            '공제총액': 0,
            '실지급액': base + meal + extra
        }


# === Excel 시트 생성 (매장별) ===

def _write_headers(ws, row: int, headers: List[str]):
    """헤더 행 작성"""
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER


def _init_totals() -> Dict:
    """합계용 딕셔너리 초기화"""
    return {k: 0 for k in [
        'base', 'meal', 'extra', 'gross', 'pension', 'health',
        'longterm', 'employ', 'income_tax', 'local_tax', 'deduction', 'net'
    ]}


def _add_to_totals(totals: Dict, calc: Dict, base: int, meal: int, extra: int):
    """합계 누적"""
    totals['base'] += base
    totals['meal'] += meal
    totals['extra'] += extra
    totals['gross'] += calc['지급총액']
    totals['pension'] += calc.get('국민연금', 0)
    totals['health'] += calc.get('건강보험', 0)
    totals['longterm'] += calc.get('장기요양', 0)
    totals['employ'] += calc.get('고용보험', 0)
    totals['income_tax'] += calc.get('소득세', 0)
    totals['local_tax'] += calc.get('지방소득세', 0)
    totals['deduction'] += calc.get('공제총액', 0)
    totals['net'] += calc.get('실지급액', calc['지급총액'])


def create_branch_sheet(wb: Workbook, branch_name: str,
                        employees: List[Dict], allowances: Dict,
                        allowance_sections: Dict, parttimer_sections: Dict,
                        year_month: str, is_first: bool = False) -> Dict:
    """매장별 시트 생성 (정직원 + 수당상세 + 파트타이머)"""
    if is_first:
        ws = wb.active
        ws.title = branch_name
    else:
        ws = wb.create_sheet(branch_name)

    # 제목
    ws.merge_cells('A1:O1')
    ws['A1'] = f"{year_month.replace('-', '년 ')}월 급여대장 - {branch_name}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws['A2'] = f"작성일: {datetime.now().strftime('%Y-%m-%d')} | 제출처: 세무사"

    # --- 정직원 섹션 ---
    row = 4
    ws.cell(row=row, column=1, value="[정직원]").font = Font(bold=True, size=11)
    for c in range(1, 16):
        ws.cell(row=row, column=c).fill = SECTION_FILL
    row += 1

    headers = ['직원명', '직급', '기본급', '식대', '추가수당', '지급총액',
               '국민연금', '건강보험', '장기요양', '고용보험', '소득세', '지방소득세',
               '공제총액', '실지급액', '비고']
    _write_headers(ws, row, headers)
    row += 1

    totals = _init_totals()
    emp_count = 0

    for emp in employees:
        name = emp.get('이름', '')
        title = emp.get('직급', '')
        base = emp['기본급']
        meal = emp['식대']

        extra = 0
        note = '-'
        if name in allowances:
            extra = allowances[name].get('amount', 0)
            note = allowances[name].get('note', '-')

        if name in allowances and 'prorated' in allowances[name]:
            base = allowances[name]['prorated_base']
            meal = allowances[name]['prorated_meal']
            note = allowances[name].get('note', '일할계산')

        calc = calculate_payroll(base, meal, extra)

        data = [
            name, title, base, meal, extra, calc['지급총액'],
            calc.get('국민연금', 0), calc.get('건강보험', 0), calc.get('장기요양', 0),
            calc.get('고용보험', 0), calc.get('소득세', 0), calc.get('지방소득세', 0),
            calc.get('공제총액', 0), calc.get('실지급액', calc['지급총액']), note
        ]

        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = THIN_BORDER
            if 3 <= col <= 14:
                cell.number_format = MONEY_FORMAT

        _add_to_totals(totals, calc, base, meal, extra)
        emp_count += 1
        row += 1

    # 정직원 소계
    if emp_count > 0:
        ws.cell(row=row, column=1, value="소계").font = Font(bold=True)
        for col, key in [(3, 'base'), (4, 'meal'), (5, 'extra'), (6, 'gross'),
                         (13, 'deduction'), (14, 'net')]:
            ws.cell(row=row, column=col, value=totals[key]).number_format = MONEY_FORMAT
        for c in range(1, 16):
            ws.cell(row=row, column=c).fill = SUBTOTAL_FILL
            ws.cell(row=row, column=c).border = THIN_BORDER
        row += 2

    # --- 수당 상세 섹션 ---
    branch_keyword = BRANCH_KEYWORDS.get(branch_name, branch_name)
    branch_allow = {k: v for k, v in allowance_sections.items() if branch_keyword in k}

    if branch_allow:
        ws.cell(row=row, column=1, value="[수당 상세]").font = Font(bold=True, size=11)
        for c in range(1, 16):
            ws.cell(row=row, column=c).fill = SECTION_FILL
        row += 1

        for section_name, items in branch_allow.items():
            display = section_name.split(' - ', 1)[1] if ' - ' in section_name else section_name
            ws.cell(row=row, column=1, value=display).font = Font(bold=True)
            row += 1

            if isinstance(items, list) and items and 'headers' in items[0]:
                for col, h in enumerate(items[0]['headers'], 1):
                    cell = ws.cell(row=row, column=col, value=h)
                    cell.font = HEADER_FONT
                    cell.fill = HEADER_FILL
                    cell.border = THIN_BORDER
                row += 1

                for item in items[1:]:
                    if 'data' in item:
                        for col, val in enumerate(item['data'], 1):
                            cell = ws.cell(row=row, column=col, value=val)
                            cell.border = THIN_BORDER
                            if isinstance(val, int) and val > 1000:
                                cell.number_format = MONEY_FORMAT
                        row += 1
            row += 1

    # --- 파트타이머 섹션 ---
    branch_pt = {k: v for k, v in parttimer_sections.items() if branch_keyword in k}

    if branch_pt:
        ws.cell(row=row, column=1, value="[파트타이머]").font = Font(bold=True, size=11)
        for c in range(1, 16):
            ws.cell(row=row, column=c).fill = SECTION_FILL
        row += 1

        for section_name, items in branch_pt.items():
            if isinstance(items, list) and items and 'headers' in items[0]:
                for col, h in enumerate(items[0]['headers'], 1):
                    cell = ws.cell(row=row, column=col, value=h)
                    cell.font = HEADER_FONT
                    cell.fill = HEADER_FILL
                    cell.border = THIN_BORDER
                row += 1

                for item in items[1:]:
                    if 'data' in item:
                        for col, val in enumerate(item['data'], 1):
                            cell = ws.cell(row=row, column=col, value=val)
                            cell.border = THIN_BORDER
                            if isinstance(val, int) and val > 1000:
                                cell.number_format = MONEY_FORMAT
                        row += 1
            row += 1

    # 열 너비
    widths = {'A': 10, 'B': 10, 'C': 12, 'D': 10, 'E': 10, 'F': 12,
              'G': 10, 'H': 10, 'I': 10, 'J': 10, 'K': 10, 'L': 10,
              'M': 12, 'N': 12, 'O': 18}
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w

    return totals


def _parse_work_days(date_str: str) -> List[int]:
    """실제 근무일 문자열에서 일(day) 숫자 리스트 추출
    예: '2/2, 2/5, 2/8' -> [2, 5, 8]
        '2, 5, 8, 13' -> [2, 5, 8, 13]
    """
    days = []
    for part in str(date_str).split(','):
        part = part.strip()
        if '/' in part:
            day = part.split('/')[-1].strip()
        else:
            day = part
        match = re.search(r'(\d+)', day)
        if match:
            days.append(int(match.group(1)))
    return sorted(days)


def create_daily_worker_sheet(wb: Workbook, daily_worker_sections: Dict,
                              year_month: str):
    """일용직 시트 생성 - 근로내용 확인신고서(별지 제22호의7서식) 양식 기반

    공식 양식 구조:
    1. 사업장 정보
    2. 근로자 인적사항 테이블 (번호/성명/주민등록번호/직종부호/매장)
    3. 근로일자 캘린더 (1~31 가로, 근로자별 행, O 표시)
    4. 근로 요약 테이블 (근로일수/일평균근로시간/보수지급기초일수/보수총액/임금총액)
    5. 국세청 일용근로소득신고 테이블
    """
    if not daily_worker_sections:
        return

    from openpyxl.utils import get_column_letter
    import calendar
    year = int(year_month[:4])
    month = int(year_month[5:7])
    days_in_month = calendar.monthrange(year, month)[1]

    ws = wb.create_sheet("일용직")

    LABEL_FILL = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
    CAL_HEADER_FILL = PatternFill(start_color='B4C6E7', end_color='B4C6E7', fill_type='solid')
    CAL_O_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    center = Alignment(horizontal='center', vertical='center')
    left_wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # --- 모든 근로자 데이터 먼저 수집 ---
    workers = []
    for section_name, items in daily_worker_sections.items():
        branch = section_name.split(' - ')[0] if ' - ' in section_name else section_name
        for item in items:
            if 'data' not in item:
                continue
            d = item['data']
            name = d[0]
            jumin = d[1] if d[1] != '-' else '(주민번호 필요)'
            pay = d[2] if isinstance(d[2], int) else parse_money(str(d[2]))
            work_dates_str = d[3] if len(d) > 3 else ''
            work_days_count_str = d[4] if len(d) > 4 else ''
            work_days = _parse_work_days(work_dates_str)
            work_count = len(work_days) if work_days else (
                int(re.search(r'(\d+)', str(work_days_count_str)).group(1))
                if re.search(r'(\d+)', str(work_days_count_str)) else 0
            )
            avg_hours = round(pay / (work_count * 11000), 1) if work_count > 0 and pay > 0 else 0
            # 소득세 계산
            daily_wage = pay / work_count if work_count > 0 else 0
            taxable_per_day = max(0, daily_wage - 150000)
            income_tax_per_day = int(taxable_per_day * 0.06 * 0.45)
            total_income_tax = income_tax_per_day * work_count
            if total_income_tax < 1000:
                total_income_tax = 0
            local_tax = int(total_income_tax * 0.1)
            workers.append({
                'branch': branch, 'name': name, 'jumin': jumin,
                'pay': pay, 'work_days': work_days, 'work_count': work_count,
                'avg_hours': avg_hours, 'income_tax': total_income_tax,
                'local_tax': local_tax, 'daily_wage': daily_wage,
            })

    if not workers:
        return

    # ===========================================
    # 1. 제목
    # ===========================================
    last_col = 2 + days_in_month  # B~(B+31) for calendar
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    ws['A1'] = f"[고용보험/산재보험] 근로내용 확인신고서(일용근로자) ({year}년 {month}월분)"
    ws['A1'].font = Font(bold=True, size=13)
    ws['A1'].alignment = center

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
    ws['A2'] = f"작성일: {datetime.now().strftime('%Y-%m-%d')} | 별지 제22호의7서식 기반 세무사 전달용"
    ws['A2'].alignment = Alignment(horizontal='right')
    ws['A2'].font = Font(size=9, color='666666')

    # ===========================================
    # 헬퍼: 셀 병합 + 값/스타일 적용
    # ===========================================
    def _merged_cell(ws, row, col_start, col_end, value,
                     font=None, fill=None, alignment=center, border=THIN_BORDER,
                     number_format=None):
        """셀 병합 후 값과 스타일 적용"""
        if col_start < col_end:
            ws.merge_cells(start_row=row, start_column=col_start,
                           end_row=row, end_column=col_end)
        cell = ws.cell(row=row, column=col_start, value=value)
        if font:
            cell.font = font
        if fill:
            cell.fill = fill
        if alignment:
            cell.alignment = alignment
        if border:
            cell.border = border
            # 병합된 셀의 오른쪽 끝에도 border
            if col_start < col_end:
                ws.cell(row=row, column=col_end).border = border
        if number_format:
            cell.number_format = number_format
        return cell

    # 테이블 열 배치 (병합 범위) - 캘린더 좁은 열에서도 보이도록
    # 인적사항: 번호(A) | 성명(B-D) | 주민등록번호(E-I) | 직종부호(J-N) | 매장(O-S)
    INFO_COLS = [(1, 1), (2, 4), (5, 9), (10, 14), (15, 19)]
    # 요약: 성명(A) | 근로일수(B-D) | 일평균근로시간(E-G) | 보수지급기초일수(H-K) | 보수총액(L-O) | 임금총액(P-S)
    SUMMARY_COLS = [(1, 1), (2, 4), (5, 7), (8, 11), (12, 15), (16, 19)]
    # 국세청: 성명(A) | 지급월(B-D) | 총지급액(E-H) | 비과세소득(I-K) | 소득세(L-O) | 지방소득세(P-S)
    TAX_COLS = [(1, 1), (2, 4), (5, 8), (9, 11), (12, 15), (16, 19)]

    # ===========================================
    # 2. 근로자 인적사항 테이블
    # ===========================================
    row = 4
    info_headers = ['번호', '성명', '주민등록번호', '직종부호', '매장']
    for (cs, ce), h in zip(INFO_COLS, info_headers):
        _merged_cell(ws, row, cs, ce, h,
                     font=HEADER_FONT, fill=HEADER_FILL)

    for idx, w in enumerate(workers, 1):
        row += 1
        vals = [idx, w['name'], w['jumin'], '532 (식당 서비스원)', w['branch']]
        for (cs, ce), v in zip(INFO_COLS, vals):
            f = Font(bold=True) if v == w['name'] else None
            _merged_cell(ws, row, cs, ce, v, font=f)

    # ===========================================
    # 3. 근로일자 캘린더 (1~31 가로, 근로자별 행)
    # ===========================================
    row += 2
    cal_start_row = row

    # 헤더: A열 = "근로일자", B~AF열 = 1~31
    cell = ws.cell(row=row, column=1, value="근로일자")
    cell.font = Font(bold=True, size=11)
    cell.fill = LABEL_FILL
    cell.border = THIN_BORDER
    cell.alignment = center

    for day in range(1, days_in_month + 1):
        col = 1 + day  # B=2 -> day 1, C=3 -> day 2, ...
        cell = ws.cell(row=row, column=col, value=day)
        cell.font = Font(bold=True, size=9)
        cell.fill = CAL_HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = center

    # 근로자별 행
    for w in workers:
        row += 1
        cell = ws.cell(row=row, column=1, value=w['name'])
        cell.font = Font(bold=True)
        cell.border = THIN_BORDER
        cell.alignment = center

        for day in range(1, days_in_month + 1):
            col = 1 + day
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            cell.alignment = center
            if day in w['work_days']:
                cell.value = "O"
                cell.fill = CAL_O_FILL
                cell.font = Font(bold=True, color='CC0000')

    # ===========================================
    # 4. 근로 요약 테이블
    # ===========================================
    row += 2
    summary_headers = ['성명', '근로일수', '일평균근로시간', '보수지급기초일수',
                        '보수총액', '임금총액']
    for (cs, ce), h in zip(SUMMARY_COLS, summary_headers):
        _merged_cell(ws, row, cs, ce, h,
                     font=HEADER_FONT, fill=HEADER_FILL)

    for w in workers:
        row += 1
        vals = [
            w['name'],
            f"{w['work_count']}일",
            f"{w['avg_hours']}시간",
            f"{w['work_count']}일",
            w['pay'],
            w['pay'],
        ]
        for i, ((cs, ce), v) in enumerate(zip(SUMMARY_COLS, vals)):
            nf = MONEY_FORMAT if i >= 4 and isinstance(v, int) else None
            _merged_cell(ws, row, cs, ce, v, number_format=nf)

    # ===========================================
    # 5. 국세청 일용근로소득신고 테이블
    # ===========================================
    row += 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=19)
    ws.cell(row=row, column=1, value="[국세청 일용근로소득신고]").font = Font(bold=True, size=11)
    row += 1

    tax_headers = ['성명', '지급월', '총지급액(과세소득)', '비과세소득', '소득세', '지방소득세']
    for (cs, ce), h in zip(TAX_COLS, tax_headers):
        _merged_cell(ws, row, cs, ce, h,
                     font=HEADER_FONT, fill=HEADER_FILL)

    for w in workers:
        row += 1
        vals = [w['name'], f"{month}월", w['pay'], 0, w['income_tax'], w['local_tax']]
        for i, ((cs, ce), v) in enumerate(zip(TAX_COLS, vals)):
            nf = MONEY_FORMAT if i >= 2 and isinstance(v, int) else None
            _merged_cell(ws, row, cs, ce, v, number_format=nf)

    # ===========================================
    # 6. 소득세 계산 근거 (참고)
    # ===========================================
    row += 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=19)
    ws.cell(row=row, column=1, value="[소득세 계산 근거]").font = Font(bold=True, size=9, color='666666')
    row += 1
    for w in workers:
        note = (f"{w['name']}: 일급 {w['daily_wage']:,.0f}원"
                f" - 근로소득공제 150,000원"
                f" = 과세대상 {max(0, w['daily_wage'] - 150000):,.0f}원"
                f" x 6% x (1-55%세액공제)"
                f" = 일 소득세 {int(max(0, w['daily_wage'] - 150000) * 0.06 * 0.45):,}원"
                f" x {w['work_count']}일"
                f" = {w['income_tax']:,}원")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
        ws.cell(row=row, column=1, value=note).font = Font(size=8, color='888888')
        row += 1

    # ===========================================
    # 열 너비 설정
    # ===========================================
    ws.column_dimensions['A'].width = 14
    # 캘린더 날짜 열 (day 1~31: col B~AF)
    for day in range(1, days_in_month + 1):
        col_letter = get_column_letter(1 + day)
        ws.column_dimensions[col_letter].width = 4.5

    print(f"  일용직: {len(workers)}명")


def create_summary_sheet(wb: Workbook, branch_totals: Dict[str, Dict],
                         year_month: str) -> Dict:
    """전체_합계 시트 생성"""
    ws = wb.create_sheet("전체_합계")

    ws.merge_cells('A1:M1')
    ws['A1'] = f"{year_month.replace('-', '년 ')}월 급여대장 - 전체 합계"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws['A2'] = f"작성일: {datetime.now().strftime('%Y-%m-%d')} | 제출처: 세무사"

    headers = ['매장', '기본급', '식대', '추가수당', '지급총액',
               '국민연금', '건강보험', '장기요양', '고용보험', '소득세', '지방소득세',
               '공제총액', '실지급액']
    _write_headers(ws, 4, headers)

    row = 5
    grand = _init_totals()
    total_keys = ['base', 'meal', 'extra', 'gross', 'pension', 'health',
                  'longterm', 'employ', 'income_tax', 'local_tax', 'deduction', 'net']

    for branch_name, totals in branch_totals.items():
        data = [branch_name] + [totals.get(k, 0) for k in total_keys]

        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = THIN_BORDER
            if col >= 2:
                cell.number_format = MONEY_FORMAT

        for k in total_keys:
            grand[k] += totals.get(k, 0)

        row += 1

    # 전체 합계 행
    row += 1
    ws.cell(row=row, column=1, value="전체 합계").font = Font(bold=True)
    for i, k in enumerate(total_keys):
        ws.cell(row=row, column=i + 2, value=grand[k]).number_format = MONEY_FORMAT

    for c in range(1, 14):
        ws.cell(row=row, column=c).fill = TOTAL_FILL
        ws.cell(row=row, column=c).border = THIN_BORDER

    widths = {'A': 12, 'B': 12, 'C': 10, 'D': 10, 'E': 12,
              'F': 10, 'G': 10, 'H': 10, 'I': 10, 'J': 10, 'K': 10,
              'L': 12, 'M': 12}
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w

    return grand


# === 메인 함수 ===

def generate_payroll_excel(year_month: str, preview: bool = False) -> Optional[str]:
    """급여대장 Excel 생성 (매장별 시트)"""
    print(f"\n=== {year_month} 급여대장 생성 (매장별 시트) ===\n")

    config = load_config()
    service = get_sheets_service()
    spreadsheet_id = config['google_sheets']['spreadsheet_id']

    employees = read_employees(service, spreadsheet_id)
    if not employees:
        print("Error: 직원 정보가 없습니다.")
        return None

    print(f"직원 수: {len(employees)}명")

    allowances, allowance_sections, parttimer_sections, daily_worker_sections = parse_payroll_changes(year_month)

    # 매장별 그룹핑
    branch_employees = {}
    for emp in employees:
        branch = emp.get('매장', '기타')
        if branch not in branch_employees:
            branch_employees[branch] = []
        branch_employees[branch].append(emp)

    if preview:
        print("\n[미리보기 모드 - 매장별]")
        for branch in BRANCH_ORDER:
            emps = branch_employees.get(branch, [])
            if not emps:
                continue
            print(f"\n--- {branch} ({len(emps)}명) ---")
            for emp in emps:
                name = emp.get('이름', '')
                base = emp['기본급']
                meal = emp['식대']
                extra_info = ""
                if name in allowances:
                    al = allowances[name]
                    if al.get('prorated'):
                        extra_info = f" [{al['note']}]"
                    elif al.get('amount', 0) > 0:
                        extra_info = f" +수당 {al['amount']:,}원 ({al['note']})"
                print(f"  {name}: {base:,}원 + {meal:,}원{extra_info}")

            branch_keyword = BRANCH_KEYWORDS.get(branch, branch)
            pt = {k: v for k, v in parttimer_sections.items() if branch_keyword in k}
            if pt:
                for section_name, items in pt.items():
                    pt_count = len([i for i in items if 'data' in i]) if items else 0
                    print(f"  [파트] {pt_count}명")

        if daily_worker_sections:
            print(f"\n--- 일용직 (전 매장 통합) ---")
            for section_name, items in daily_worker_sections.items():
                dw_count = len([i for i in items if 'data' in i]) if items else 0
                print(f"  {section_name}: {dw_count}명")

        return None

    # Excel 생성
    wb = Workbook()
    branch_totals = {}
    is_first = True

    for branch in BRANCH_ORDER:
        emps = branch_employees.get(branch, [])
        if not emps:
            continue

        totals = create_branch_sheet(
            wb, branch, emps, allowances,
            allowance_sections, parttimer_sections,
            year_month, is_first=is_first
        )
        branch_totals[branch] = totals
        is_first = False
        print(f"  {branch}: {len(emps)}명, 지급총액 {totals['gross']:,}원")

    # BRANCH_ORDER에 없는 매장
    for branch, emps in branch_employees.items():
        if branch not in BRANCH_ORDER and emps:
            totals = create_branch_sheet(
                wb, branch, emps, allowances,
                allowance_sections, parttimer_sections,
                year_month, is_first=is_first
            )
            branch_totals[branch] = totals
            is_first = False

    # 일용직 시트 (있으면 생성)
    create_daily_worker_sheet(wb, daily_worker_sections, year_month)

    # 전체_합계 시트
    grand = create_summary_sheet(wb, branch_totals, year_month)

    # 저장
    year = year_month[:4]
    output_dir = resolve_path(_CFG, 'excel_folder', year)
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"{year_month}_급여대장.xlsx"

    wb.save(output_path)

    print(f"\n저장: {output_path}")

    # Google Sheets에도 저장 (퇴직금 계산용)
    print(f"\n=== Google Sheets 저장 ===")
    try:
        save_to_google_sheets(service, spreadsheet_id, year_month,
                              employees, allowances, branch_employees)
    except Exception as e:
        print(f"  Warning: Google Sheets 저장 실패: {e}")

    print(f"\n=== 요약 ===")
    print(f"시트: {', '.join(branch_totals.keys())}, 전체_합계")
    print(f"정직원 지급총액: {grand['gross']:,}원")
    print(f"정직원 공제총액: {grand['deduction']:,}원")
    print(f"정직원 실지급액: {grand['net']:,}원")

    return str(output_path)


def save_to_google_sheets(service, spreadsheet_id: str, year_month: str,
                          employees: List[Dict], allowances: Dict,
                          branch_employees: Dict):
    """Google Sheets에 급여대장 시트 저장 (퇴직금 계산용)

    YYYY-MM_급여대장 시트에 전 직원 급여 데이터 저장.
    retirement.py가 최근 3개월 급여를 조회할 때 사용.
    """
    sheet_name = f"{year_month}_급여대장"

    # 시트 존재 여부 확인
    sheet_metadata = service.spreadsheets().get(
        spreadsheetId=spreadsheet_id).execute()
    existing_sheets = [s['properties']['title']
                       for s in sheet_metadata.get('sheets', [])]

    if sheet_name not in existing_sheets:
        # 시트 생성
        service.spreadsheets().batchUpdate(
            spreadsheetId=spreadsheet_id,
            body={'requests': [{'addSheet': {'properties': {'title': sheet_name}}}]}
        ).execute()
        print(f"  시트 생성: {sheet_name}")
    else:
        print(f"  시트 존재: {sheet_name} (덮어쓰기)")

    # 데이터 준비
    headers = ['이름', '매장', '기본급', '식대', '추가수당', '지급총액']
    rows = [headers]

    for branch in BRANCH_ORDER:
        emps = branch_employees.get(branch, [])
        for emp in emps:
            name = emp.get('이름', '')
            base = emp['기본급']
            meal = emp['식대']
            extra = 0
            if name in allowances:
                al = allowances[name]
                if al.get('prorated'):
                    # 일할계산: base를 일할 금액으로 대체
                    base = al.get('prorated_base', base)
                    extra = al.get('amount', 0)
                else:
                    extra = al.get('amount', 0)
            gross = base + meal + extra
            rows.append([name, branch, base, meal, extra, gross])

    # 시트에 쓰기
    service.spreadsheets().values().update(
        spreadsheetId=spreadsheet_id,
        range=f"'{sheet_name}'!A1",
        valueInputOption='RAW',
        body={'values': rows}
    ).execute()

    print(f"  저장 완료: {len(rows) - 1}명 -> {sheet_name}")
    return sheet_name


def main():
    parser = argparse.ArgumentParser(description='급여대장 Excel 생성 (매장별 시트)')
    parser.add_argument('year_month', help='대상 월 (YYYY-MM)')
    parser.add_argument('--preview', action='store_true', help='미리보기만')

    args = parser.parse_args()

    generate_payroll_excel(args.year_month, args.preview)


if __name__ == '__main__':
    main()
